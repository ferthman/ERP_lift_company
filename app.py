from flask import Flask, request, jsonify, render_template, send_from_directory
from flask_cors import CORS
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, func
from sqlalchemy.orm import declarative_base, sessionmaker, relationship, scoped_session
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import sys
from datetime import datetime, timezone

# ---------------------------------------------------------------------------
# Vendor path configuration
#
# To reduce external dependencies, we include a vendored copy of the
# openpyxl library in the `vendor/` directory. If openpyxl is not
# installed in the Python environment, adding this directory to sys.path
# allows the vendored version to be imported transparently.
vendor_dir = os.path.join(os.path.dirname(__file__), "vendor")
if os.path.isdir(vendor_dir) and vendor_dir not in sys.path:
    sys.path.append(vendor_dir)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret")
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False
CORS(app, supports_credentials=True)

login_manager = LoginManager(app)
login_manager.login_view = "index"

DB_PATH = os.path.join(os.path.dirname(__file__), "lift_crm.db")
engine = create_engine(f"sqlite:///{DB_PATH}", echo=False, future=True)
SessionLocal = scoped_session(sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False))
Base = declarative_base()

class Master(Base):
    __tablename__ = "masters"
    id = Column(Integer, primary_key=True)
    name = Column(String, nullable=False)
    is_active = Column(Integer, default=1)
    user = relationship("User", uselist=False, back_populates="master")
    tickets = relationship("Ticket", back_populates="assigned_master")

class User(Base, UserMixin):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True)
    username = Column(String, unique=True, nullable=False)
    password_hash = Column(String, nullable=False)
    role = Column(String, nullable=False)  # admin | dispatcher | master
    master_id = Column(Integer, ForeignKey("masters.id"), nullable=True)
    master = relationship("Master", back_populates="user")
    def get_id(self): return str(self.id)

class Ticket(Base):
    __tablename__ = "tickets"
    id = Column(Integer, primary_key=True)
    object_name = Column(String, nullable=False)
    address = Column(String, nullable=True)
    lat = Column(Float, nullable=False)
    lon = Column(Float, nullable=False)
    description = Column(String, nullable=True)
    status = Column(String, default="NEW")
    assigned_master_id = Column(Integer, ForeignKey("masters.id"), nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    arrived_at = Column(DateTime, nullable=True)
    completed_at = Column(DateTime, nullable=True)
    arrival_lat = Column(Float, nullable=True)
    arrival_lon = Column(Float, nullable=True)
    completion_lat = Column(Float, nullable=True)
    completion_lon = Column(Float, nullable=True)
    assigned_master = relationship("Master", back_populates="tickets")
    attachments = relationship("Attachment", back_populates="ticket", cascade="all, delete-orphan")

    # Email address of the customer who created the ticket. This field is optional.
    # It allows the system to send an automatic report to the customer when the
    # ticket is completed. If None, no report will be emailed.
    email = Column(String, nullable=True)

class Attachment(Base):
    __tablename__ = "attachments"
    id = Column(Integer, primary_key=True)
    ticket_id = Column(Integer, ForeignKey("tickets.id"), nullable=False)
    filename = Column(String, nullable=False)
    orig_name = Column(String, nullable=False)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    ticket = relationship("Ticket", back_populates="attachments")

def init_db():
    Base.metadata.create_all(engine)
    with SessionLocal() as db:
        if db.query(Master).count() == 0:
            masters = [Master(name=f"Мастер #{i+1}") for i in range(10)]
            db.add_all(masters); db.commit()
        if db.query(User).count() == 0:
            admin = User(username="admin", password_hash=generate_password_hash("admin123"), role="admin")
            disp = User(username="dispatcher", password_hash=generate_password_hash("disp123"), role="dispatcher")
            db.add_all([admin, disp]); db.commit()
            for m in db.query(Master).order_by(Master.id).all():
                u = User(username=f"master{m.id}", password_hash=generate_password_hash("m123456"), role="master", master_id=m.id)
                db.add(u)
            db.commit()

def ensure_migrations():
    try:
        import sqlite3
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        # Check for the 'is_active' column on the masters table
        cur.execute("PRAGMA table_info(masters)")
        cols = [r[1] for r in cur.fetchall()]
        if 'is_active' not in cols:
            cur.execute("ALTER TABLE masters ADD COLUMN is_active INTEGER DEFAULT 1")
            conn.commit()
        # Check if the tickets table has an email column and add one if missing. This
        # migration allows storing a customer email so that completion reports can
        # be sent automatically.
        cur.execute("PRAGMA table_info(tickets)")
        tcols = [r[1] for r in cur.fetchall()]
        if 'email' not in tcols:
            cur.execute("ALTER TABLE tickets ADD COLUMN email TEXT")
            conn.commit()
        conn.close()
    except Exception as e:
        print("Migration check failed:", e)

@login_manager.user_loader
def load_user(user_id):
    with SessionLocal() as db:
        return db.get(User, int(user_id))

def role_required(*roles):
    def decorator(fn):
        from functools import wraps
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated: return jsonify({"error":"Unauthorized"}), 401
            if current_user.role not in roles: return jsonify({"error":"Forbidden"}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

def _to_utc(dt):
    from datetime import timezone
    if dt is None:
        return None
    return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt.astimezone(timezone.utc)


def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000.0
    from math import radians, sin, cos, atan2, sqrt
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1); dl = radians(lon2 - lon1)
    a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dl/2)**2
    c = 2*atan2(sqrt(a), sqrt(1-a))
    return R * c

def auto_assign_master(db):
    open_statuses = ["NEW","ASSIGNED","IN_PROGRESS"]
    active = db.query(Master).filter(Master.is_active==1).all()
    counts = {m.id:0 for m in active}
    if active:
        rows = db.query(Ticket.assigned_master_id, func.count(Ticket.id)).filter(
            Ticket.status.in_(open_statuses),
            Ticket.assigned_master_id.in_([m.id for m in active])
        ).group_by(Ticket.assigned_master_id).all()
        for mid, cnt in rows: counts[mid]=cnt
    if not counts: return None
    best_id = min(counts, key=lambda k: (counts[k], k))
    return db.get(Master, best_id)

# ---------------------------------------------------------------------------
# Reporting and Archiving Helpers
#
# The following helpers implement optional email notifications and archiving
# functionality. When a ticket is completed, a report can be sent to the
# customer via SMTP if an email address is provided. When a ticket is deleted,
# its details are archived into an Excel file for future reference.

def send_report(ticket: Ticket):
    """
    Attempt to send a completion report to the customer's email address. This
    function checks environment variables for SMTP configuration. If the
    configuration is incomplete or the ticket has no email, the function
    silently returns. Any exceptions during sending are printed to stderr.

    The following environment variables are used:
      SMTP_SERVER: hostname of the SMTP server (required)
      SMTP_PORT: port of the SMTP server (required)
      SMTP_USERNAME: username for authentication (optional)
      SMTP_PASSWORD: password for authentication (optional)
    """
    # Only send a report if the ticket has an email associated
    if not getattr(ticket, "email", None):
        return
    import smtplib
    from email.mime.text import MIMEText

    smtp_server = os.environ.get("SMTP_SERVER")
    smtp_port = os.environ.get("SMTP_PORT")
    smtp_user = os.environ.get("SMTP_USERNAME")
    smtp_pass = os.environ.get("SMTP_PASSWORD")
    if not smtp_server or not smtp_port:
        # Missing configuration; skip sending
        print("Email report skipped: SMTP_SERVER or SMTP_PORT not configured")
        return
    # Compose the email
    subject = f"Заявка {ticket.id} — {ticket.object_name} завершена"
    body = (
        f"Здравствуйте!\n\n"
        f"Ваша заявка '{ticket.object_name}' была завершена.\n"
        f"Статус: {ticket.status}\n"
        f"Описание: {ticket.description or ''}\n"
        f"Адрес: {ticket.address or ''}\n\n"
        f"Спасибо за обращение."
    )
    msg = MIMEText(body, _charset="utf-8")
    msg["Subject"] = subject
    msg["From"] = smtp_user or "noreply@example.com"
    msg["To"] = ticket.email
    try:
        with smtplib.SMTP(smtp_server, int(smtp_port)) as server:
            server.starttls()
            if smtp_user:
                try:
                    server.login(smtp_user, smtp_pass or "")
                except Exception:
                    # Some servers do not require login or may not support it
                    pass
            server.sendmail(msg["From"], [ticket.email], msg.as_string())
    except Exception as e:
        print("Failed to send completion report:", e)


def archive_ticket(ticket: Ticket):
    """
    Append a ticket's details to an archive Excel file. The archive file is
    stored at the project root under 'archive.xlsx'. If the file does not
    exist it will be created with headers. This function uses openpyxl to
    manipulate the Excel file. Any exceptions will be printed but will not
    interrupt the deletion flow.
    """
    from datetime import datetime
    from openpyxl import Workbook, load_workbook
    path = os.path.join(os.path.dirname(__file__), "archive.xlsx")
    # Create file with headers if it does not exist
    header = [
        "id", "object_name", "address", "lat", "lon", "description",
        "email", "status", "assigned_master_id", "assigned_master_name",
        "created_at", "updated_at", "arrived_at", "completed_at"
    ]
    try:
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            wb.save(path)
        # Load the workbook and append the new row
        wb = load_workbook(path)
        ws = wb.active
        row_data = [
            ticket.id,
            ticket.object_name,
            ticket.address,
            ticket.lat,
            ticket.lon,
            ticket.description,
            ticket.email,
            ticket.status,
            ticket.assigned_master_id,
            ticket.assigned_master.name if ticket.assigned_master else None,
            ticket.created_at.isoformat() if ticket.created_at else None,
            ticket.updated_at.isoformat() if ticket.updated_at else None,
            ticket.arrived_at.isoformat() if ticket.arrived_at else None,
            ticket.completed_at.isoformat() if ticket.completed_at else None,
        ]
        ws.append(row_data)
        wb.save(path)
        # Also save a numbered copy of the archive for each deletion. The file
        # name will be archive_N.xlsx where N is a sequential counter based on
        # existing numbered archives in the project directory. This provides a
        # snapshot of the archive at the time of deletion.
        try:
            import re, glob
            dirpath = os.path.dirname(__file__)
            existing = glob.glob(os.path.join(dirpath, "archive_*.xlsx"))
            nums = []
            for f in existing:
                m = re.search(r"archive_(\d+)\.xlsx", os.path.basename(f))
                if m:
                    nums.append(int(m.group(1)))
            next_num = max(nums) + 1 if nums else 1
            numbered_path = os.path.join(dirpath, f"archive_{next_num}.xlsx")
            import shutil
            shutil.copy2(path, numbered_path)
        except Exception as e_inner:
            print("Failed to create numbered archive copy:", e_inner)
    except Exception as e:
        print("Failed to archive ticket:", e)

ALLOWED_EXTS = {"png","jpg","jpeg","webp"}
def allowed_file(fname): return "." in fname and fname.rsplit(".",1)[-1].lower() in ALLOWED_EXTS

@app.get("/")
def index(): return render_template("index.html")

# Flag to ensure database initialization only happens once
_db_initialized = False

@app.before_request
def first_request_setup():
    global _db_initialized
    if not _db_initialized:
        init_db()
        ensure_migrations()
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
        # Ensure the objects folder exists and create default objects files if missing.
        objects_dir = os.path.join(os.path.dirname(__file__), "objects")
        os.makedirs(objects_dir, exist_ok=True)
        xlsx_path = os.path.join(objects_dir, "objects.xlsx")
        json_path = os.path.join(objects_dir, "objects.json")
        # Define a sample object for initialization
        sample_object = {
            "object_name": "Central Almaty",
            "address": "ул. Кабанбай Батыра 123",
            "lat": 43.238949,
            "lon": 76.889709
        }
        try:
            # Create Excel file if it doesn't exist
            if not os.path.exists(xlsx_path):
                try:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["object_name", "address", "lat", "lon"])
                    ws.append([sample_object["object_name"], sample_object["address"], sample_object["lat"], sample_object["lon"]])
                    wb.save(xlsx_path)
                except Exception as e:
                    # Fallback: if openpyxl fails, skip creating xlsx
                    print("Failed to create objects.xlsx:", e)
            # Create JSON file if it doesn't exist
            if not os.path.exists(json_path):
                import json
                with open(json_path, "w", encoding="utf-8") as jf:
                    json.dump([sample_object], jf, ensure_ascii=False, indent=2)
        except Exception as e:
            print("Failed to initialize objects files:", e)

        # Ensure an empty archive file exists so that the download endpoint
        # always has something to serve. Use openpyxl to create an empty file
        # with headers if it does not exist.
        archive_path = os.path.join(os.path.dirname(__file__), "archive.xlsx")
        try:
            if not os.path.exists(archive_path):
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append([
                    "id", "object_name", "address", "lat", "lon", "description",
                    "email", "status", "assigned_master_id", "assigned_master_name",
                    "created_at", "updated_at", "arrived_at", "completed_at"
                ])
                wb.save(archive_path)
        except Exception as e:
            print("Failed to initialize archive file:", e)
        _db_initialized = True


@app.post("/api/login")
def api_login():
    data = request.get_json() or {}
    username = data.get("username","").strip(); password = data.get("password","")
    with SessionLocal() as db:
        user = db.query(User).filter_by(username=username).first()
        if not user or not check_password_hash(user.password_hash, password):
            return jsonify({"error":"Неверный логин или пароль"}), 400
    login_user(user)
    return jsonify({"ok":True, "role": user.role, "username": user.username, "master_id": user.master_id})

@app.post("/api/logout")
@login_required
def api_logout(): logout_user(); return jsonify({"ok":True})

@app.get("/api/me")
def api_me():
    if not current_user.is_authenticated: return jsonify({"authenticated": False})
    return jsonify({"authenticated": True, "username": current_user.username, "role": current_user.role, "master_id": current_user.master_id})

@app.get("/api/masters")
@login_required
def list_masters():
    with SessionLocal() as db:
        ms = db.query(Master).order_by(Master.id).all()
        return jsonify([{"id": m.id, "name": m.name, "is_active": bool(m.is_active), "username": m.user.username if m.user else None} for m in ms])

@app.post("/api/masters")
@login_required
@role_required("admin")
def create_master():
    data = request.get_json() or {}; name=(data.get("name") or "").strip()
    if not name: return jsonify({"error":"Name is required"}), 400
    with SessionLocal() as db:
        m = Master(name=name, is_active=1); db.add(m); db.commit(); db.refresh(m)
        u = User(username=f"master{m.id}", password_hash=generate_password_hash("m123456"), role="master", master_id=m.id)
        db.add(u); db.commit()
        return jsonify({"id":m.id,"name":m.name,"username":u.username,"temp_password":"m123456"}), 201

@app.delete("/api/masters/<int:master_id>")
@login_required
@role_required("admin")
def delete_master(master_id):
    with SessionLocal() as db:
        m = db.get(Master, master_id)
        if not m: return jsonify({"error":"Master not found"}), 404
        others = db.query(Master).filter(Master.id!=master_id, Master.is_active==1).all()
        if not others: return jsonify({"error":"Нельзя удалить единственного активного мастера"}), 400
        open_statuses = ["NEW","ASSIGNED","IN_PROGRESS"]
        open_tickets = db.query(Ticket).filter(Ticket.assigned_master_id==master_id, Ticket.status.in_(open_statuses)).all()
        counts = {x.id:0 for x in others}
        rows = db.query(Ticket.assigned_master_id, func.count(Ticket.id)).filter(Ticket.status.in_(open_statuses), Ticket.assigned_master_id.in_([x.id for x in others])).group_by(Ticket.assigned_master_id).all()
        for mid,cnt in rows: counts[mid]=cnt
        for t in open_tickets:
            new_id = min(counts, key=lambda k:(counts[k],k))
            t.assigned_master_id = new_id
            if t.status in ["NEW","ASSIGNED"]: t.status="ASSIGNED"
            counts[new_id]+=1
        for u in db.query(User).filter(User.master_id==master_id).all(): db.delete(u)
        db.delete(m); db.commit(); return jsonify({"ok":True,"reassigned":len(open_tickets)})

@app.patch("/api/masters/<int:master_id>/toggle_active")
@login_required
@role_required("admin")
def toggle_master_active(master_id):
    with SessionLocal() as db:
        m = db.get(Master, master_id)
        if not m: return jsonify({"error":"Master not found"}), 404
        m.is_active = 0 if m.is_active else 1; reassigned = 0
        if m.is_active == 0:
            others = db.query(Master).filter(Master.id!=master_id, Master.is_active==1).all()
            if not others: return jsonify({"error":"Нет других активных мастеров для перераспределения"}), 400
            open_statuses = ["NEW","ASSIGNED","IN_PROGRESS"]
            open_tickets = db.query(Ticket).filter(Ticket.assigned_master_id==master_id, Ticket.status.in_(open_statuses)).all()
            counts = {x.id:0 for x in others}
            rows = db.query(Ticket.assigned_master_id, func.count(Ticket.id)).filter(Ticket.status.in_(open_statuses), Ticket.assigned_master_id.in_([x.id for x in others])).group_by(Ticket.assigned_master_id).all()
            for mid,cnt in rows: counts[mid]=cnt
            for t in open_tickets:
                new_id = min(counts, key=lambda k:(counts[k],k))
                t.assigned_master_id = new_id
                if t.status in ["NEW","ASSIGNED"]: t.status="ASSIGNED"
                counts[new_id]+=1
            reassigned = len(open_tickets)
        db.commit(); return jsonify({"ok":True,"is_active":bool(m.is_active),"reassigned":reassigned})

@app.get("/api/tickets")
@login_required
def list_tickets():
    with SessionLocal() as db:
        tickets = db.query(Ticket).order_by(Ticket.created_at.desc()).all()
        def ser(t: Ticket):
            return {
                "id": t.id,
                "object_name": t.object_name,
                "address": t.address,
                "lat": t.lat,
                "lon": t.lon,
                "description": t.description,
                # Include the customer's email so the front-end can show it or use it later
                "email": t.email,
                "status": t.status,
                "assigned_master_id": t.assigned_master_id,
                "assigned_master_name": t.assigned_master.name if t.assigned_master else None,
                "created_at": (_to_utc(t.created_at).isoformat() if t.created_at else None),
                "updated_at": (_to_utc(t.updated_at).isoformat() if t.updated_at else None),
                "arrived_at": (_to_utc(t.arrived_at).isoformat() if t.arrived_at else None),
                "completed_at": (_to_utc(t.completed_at).isoformat() if t.completed_at else None),
                "attachments": [
                    {"id": a.id, "url": f"/uploads/{a.filename}", "name": a.orig_name} for a in t.attachments
                ],
                "created_ts": (int(_to_utc(t.created_at).timestamp() * 1000) if t.created_at else None),
                "elapsed_ms": (
                    int(
                        ((datetime.now(timezone.utc) - _to_utc(t.created_at)).total_seconds()) * 1000
                    )
                    if t.created_at
                    else 0
                ),
            }
        return jsonify([ser(t) for t in tickets])

@app.post("/api/tickets")
@login_required
@role_required("admin","dispatcher")
def create_ticket():
    data = request.get_json() or {}
    for k in ("object_name","lat","lon"):
        if k not in data: return jsonify({"error":f"Missing field: {k}"}), 400
    with SessionLocal() as db:
        # Create a new ticket. Include the optional email address if provided so
        # that a completion report can be emailed later. Default status is NEW.
        t = Ticket(
            object_name=data["object_name"],
            address=data.get("address"),
            lat=float(data["lat"]),
            lon=float(data["lon"]),
            description=data.get("description"),
            email=data.get("email"),
            status="NEW"
        )
        m = auto_assign_master(db)
        if m:
            t.assigned_master_id, t.status = m.id, "ASSIGNED"
        db.add(t)
        db.commit()
        return jsonify({"id": t.id, "assigned_master_id": t.assigned_master_id, "status": t.status}), 201

@app.post("/api/tickets/<int:ticket_id>/reassign")
@login_required
@role_required("admin","dispatcher")
def reassign_ticket(ticket_id):
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        m = auto_assign_master(db)
        if not m: return jsonify({"error":"No active masters available"}), 400
        t.assigned_master_id = m.id
        if t.status in ["NEW","ASSIGNED"]: t.status="ASSIGNED"
        db.commit(); return jsonify({"message":"Reassigned","assigned_master_id":t.assigned_master_id})

@app.post("/api/tickets/<int:ticket_id>/cancel")
@login_required
@role_required("admin","dispatcher")
def cancel_ticket(ticket_id):
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        if t.status in ["COMPLETED","CANCELLED"]: return jsonify({"error":"Ticket already finalized"}), 400
        t.status = "CANCELLED"; db.commit(); return jsonify({"message":"Cancelled"})

@app.get("/api/tickets/<int:ticket_id>")
@login_required
def get_ticket(ticket_id):
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        return jsonify({
            "id": t.id,
            "object_name": t.object_name,
            "address": t.address,
            "lat": t.lat,
            "lon": t.lon,
            "description": t.description,
            # Include email in single ticket view
            "email": t.email,
            "status": t.status,
            "assigned_master_id": t.assigned_master_id,
            "assigned_master_name": t.assigned_master.name if t.assigned_master else None,
            "created_at": (_to_utc(t.created_at).isoformat() if t.created_at else None),
            "updated_at": (_to_utc(t.updated_at).isoformat() if t.updated_at else None),
            "arrived_at": (_to_utc(t.arrived_at).isoformat() if t.arrived_at else None),
            "completed_at": (_to_utc(t.completed_at).isoformat() if t.completed_at else None),
            "attachments": [
                {"id": a.id, "url": f"/uploads/{a.filename}", "name": a.orig_name} for a in t.attachments
            ],
            "created_ts": (int(_to_utc(t.created_at).timestamp() * 1000) if t.created_at else None),
            "elapsed_ms": (
                int(
                    ((datetime.now(timezone.utc) - _to_utc(t.created_at)).total_seconds()) * 1000
                )
                if t.created_at
                else 0
            ),
        })

@app.post("/api/tickets/<int:ticket_id>/arrive")
@login_required
def arrive_ticket(ticket_id):
    if current_user.role != "master": return jsonify({"error":"Only master can mark arrival"}), 403
    data = request.get_json() or {}; lat=data.get("lat"); lon=data.get("lon")
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        if t.assigned_master_id != current_user.master_id: return jsonify({"error":"Ticket not assigned to you"}), 403
        if lat is None or lon is None: return jsonify({"error":"lat/lon required"}), 400
        dist = haversine_m(float(lat), float(lon), t.lat, t.lon)
        if dist > 500: return jsonify({"error": f"Outside geofence ({int(dist)} m > 500 m)"}), 403
        t.status = "IN_PROGRESS"; t.arrived_at = datetime.now(timezone.utc)
        t.arrival_lat = float(lat); t.arrival_lon = float(lon)
        db.commit(); return jsonify({"message":"Arrived","distance_m":int(dist),"status":t.status})

@app.post("/api/tickets/<int:ticket_id>/complete")
@login_required
def complete_ticket(ticket_id):
    if current_user.role != "master": return jsonify({"error":"Only master can complete"}), 403
    data = request.get_json() or {}; lat=data.get("lat"); lon=data.get("lon")
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        if t.assigned_master_id != current_user.master_id: return jsonify({"error":"Ticket not assigned to you"}), 403
        if lat is None or lon is None: return jsonify({"error":"lat/lon required"}), 400
        dist = haversine_m(float(lat), float(lon), t.lat, t.lon)
        if dist > 500: return jsonify({"error": f"Outside geofence ({int(dist)} m > 500 m)"}), 403
        t.status = "COMPLETED"
        t.completed_at = datetime.now(timezone.utc)
        t.completion_lat = float(lat)
        t.completion_lon = float(lon)
        db.commit()
        # After committing the completion, attempt to send a report to the customer
        try:
            send_report(t)
        except Exception as e:
            # Avoid failing the API call if reporting fails
            print("Report sending failed:", e)
        return jsonify({"message": "Completed", "distance_m": int(dist), "status": t.status})

@app.post("/api/tickets/<int:ticket_id>/upload")
@login_required
def upload_file(ticket_id):
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        if current_user.role == "master" and t.assigned_master_id != current_user.master_id:
            return jsonify({"error":"Ticket not assigned to you"}), 403
    if "file" not in request.files: return jsonify({"error":"No file"}), 400
    f = request.files["file"]
    if f.filename == "": return jsonify({"error":"Empty filename"}), 400
    if not ('.' in f.filename and f.filename.rsplit('.',1)[-1].lower() in {"png","jpg","jpeg","webp"}):
        return jsonify({"error":"Only png/jpg/jpeg/webp allowed"}), 400
    fname = secure_filename(f.filename)
    unique = f"{int(datetime.now(timezone.utc).timestamp())}_{fname}"
    f.save(os.path.join(app.config["UPLOAD_FOLDER"], unique))
    with SessionLocal() as db:
        a = Attachment(ticket_id=ticket_id, filename=unique, orig_name=fname); db.add(a); db.commit()
    return jsonify({"ok":True, "url": f"/uploads/{unique}", "name": fname})

@app.get("/uploads/<path:filename>")
def serve_upload(filename): return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.post("/api/tickets/<int:ticket_id>/assign/<int:master_id>")
@login_required
@role_required("admin","dispatcher")
def assign_ticket(ticket_id, master_id):
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t: return jsonify({"error":"Ticket not found"}), 404
        m = db.get(Master, master_id)
        if not m: return jsonify({"error":"Master not found"}), 404
        if int(getattr(m, "is_active", 1)) != 1:
            return jsonify({"error":"Мастер неактивен"}), 400
        t.assigned_master_id = m.id
        if t.status in ["NEW","ASSIGNED"]:
            t.status = "ASSIGNED"
        db.commit()
        return jsonify({"message":"Assigned","assigned_master_id": t.assigned_master_id, "assigned_master_name": m.name})

# Delete a ticket permanently. Only admin and dispatcher can perform this
# operation. When a ticket is deleted its attachments are removed from disk
# and the ticket details are appended to an archive Excel file for
# reporting. The endpoint returns a confirmation message on success.
@app.delete("/api/tickets/<int:ticket_id>")
@login_required
@role_required("admin","dispatcher")
def delete_ticket(ticket_id):
    with SessionLocal() as db:
        t = db.get(Ticket, ticket_id)
        if not t:
            return jsonify({"error": "Ticket not found"}), 404
        # Archive the ticket before deleting
        try:
            archive_ticket(t)
        except Exception as e:
            # Log but continue
            print("Archive failed:", e)
        # Remove associated files
        for a in list(t.attachments):
            try:
                os.remove(os.path.join(app.config["UPLOAD_FOLDER"], a.filename))
            except Exception:
                pass
        db.delete(t)
        db.commit()
        return jsonify({"message": "Deleted"})

@app.get("/api/metrics")
@login_required
@role_required("admin","dispatcher")
def metrics():
    """Return various metrics about tickets and masters.

    The endpoint returns overall ticket counts by status and duration statistics,
    as well as per-master statistics. Per-master metrics include the number of
    tickets assigned to each master, counts by status, and average/median time
    to complete tickets. This allows administrators and dispatchers to monitor
    performance and workload distribution."""
    from statistics import median
    with SessionLocal() as db:
        tickets = db.query(Ticket).all()
        counts = {"NEW":0,"ASSIGNED":0,"IN_PROGRESS":0,"COMPLETED":0,"CANCELLED":0}
        for t in tickets:
            counts[t.status] = counts.get(t.status, 0) + 1
        durs = [(_to_utc(t.completed_at) - _to_utc(t.created_at)).total_seconds() for t in tickets if t.completed_at and t.created_at]
        overall = {
            "total": len(tickets),
            "counts": counts,
            "avg_close_sec": (sum(durs) / len(durs)) if durs else None,
            "median_close_sec": (median(durs) if durs else None)
        }
        masters = db.query(Master).all()
        masters_data = []
        for m in masters:
            mtickets = [t for t in tickets if t.assigned_master_id == m.id]
            m_counts = {"NEW":0,"ASSIGNED":0,"IN_PROGRESS":0,"COMPLETED":0,"CANCELLED":0}
            mdurs = []
            for t in mtickets:
                m_counts[t.status] = m_counts.get(t.status, 0) + 1
                if t.completed_at and t.created_at:
                    mdurs.append((_to_utc(t.completed_at) - _to_utc(t.created_at)).total_seconds())
            masters_data.append({
                "id": m.id,
                "name": m.name,
                "total": len(mtickets),
                "counts": m_counts,
                "avg_close_sec": (sum(mdurs) / len(mdurs)) if mdurs else None,
                "median_close_sec": (median(mdurs) if mdurs else None)
            })
        return jsonify({"overall": overall, "masters": masters_data})
@app.get("/api/health")
def health(): return jsonify({"ok":True})

# Serve the archive Excel file for download. Only authenticated users can
# download the archive. The file is stored at the project root as
# 'archive.xlsx'. A 'download' query parameter can be used to force a
# browser download.
@app.get("/api/archive")
@login_required
@role_required("admin", "dispatcher")
def download_archive():
    archive_path = os.path.join(os.path.dirname(__file__), "archive.xlsx")
    # Generate an up-to-date Excel export containing all current tickets.
    # If openpyxl is not available, inform the client that a dependency is missing
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({"error": "Missing dependency: openpyxl. Please install requirements."}), 500
    wb = Workbook()
    ws = wb.active
    # Header row
    ws.append([
        "id", "object_name", "address", "lat", "lon", "description",
        "email", "status", "assigned_master_id", "assigned_master_name",
        "created_at", "updated_at", "arrived_at", "completed_at"
    ])
    # Populate with all tickets from the database
    with SessionLocal() as db:
        tickets = db.query(Ticket).order_by(Ticket.id).all()
        for t in tickets:
            ws.append([
                t.id,
                t.object_name,
                t.address,
                t.lat,
                t.lon,
                t.description,
                t.email,
                t.status,
                t.assigned_master_id,
                t.assigned_master.name if t.assigned_master else None,
                t.created_at.isoformat() if t.created_at else None,
                t.updated_at.isoformat() if t.updated_at else None,
                t.arrived_at.isoformat() if t.arrived_at else None,
                t.completed_at.isoformat() if t.completed_at else None,
            ])
    # Save to the archive file within the project directory
    export_path = archive_path
    try:
        wb.save(export_path)
    except Exception as e:
        print("Failed to save export:", e)
    return send_from_directory(
        directory=os.path.dirname(export_path),
        path=os.path.basename(export_path),
        as_attachment=True,
        download_name=os.path.basename(export_path)
    )

# List objects from the objects Excel file. The file is expected in
# 'objects/objects.xlsx'. Returns an array of objects with lat, lon, and
# name/address fields. If the file is missing or empty, returns an empty array.
@app.get("/api/objects")
@login_required
def api_objects():
    """
    Return objects from the objects file. Objects can be stored in either
    `objects.xlsx` (preferred) or `objects.json`. If openpyxl is unavailable
    or the Excel file cannot be read, the JSON file will be used as a
    fallback. Each object must provide lat and lon coordinates.
    """
    objects_dir = os.path.join(os.path.dirname(__file__), "objects")
    xlsx_path = os.path.join(objects_dir, "objects.xlsx")
    json_path = os.path.join(objects_dir, "objects.json")
    records = []
    # Try reading Excel file first if it exists and openpyxl is available
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h) if h is not None else "" for h in rows[0]]
                for r in rows[1:]:
                    obj = {header[i]: r[i] for i in range(len(header))}
                    try:
                        lat = float(obj.get("lat")) if obj.get("lat") is not None else None
                        lon = float(obj.get("lon")) if obj.get("lon") is not None else None
                    except Exception:
                        lat = None; lon = None
                    if lat is None or lon is None:
                        continue
                    records.append({
                        "object_name": str(obj.get("object_name", "")),
                        "address": str(obj.get("address", "")),
                        "lat": lat,
                        "lon": lon,
                    })
        except Exception as e:
            # Log and fall back to JSON
            print("Failed to read objects.xlsx:", e)
            records = []
    # If no records loaded from Excel and a JSON file exists, load from JSON
    if not records and os.path.exists(json_path):
        try:
            import json
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                for obj in data:
                    try:
                        lat = float(obj.get("lat")) if obj.get("lat") is not None else None
                        lon = float(obj.get("lon")) if obj.get("lon") is not None else None
                    except Exception:
                        lat = None; lon = None
                    if lat is None or lon is None:
                        continue
                    records.append({
                        "object_name": str(obj.get("object_name", "")),
                        "address": str(obj.get("address", "")),
                        "lat": lat,
                        "lon": lon,
                    })
        except Exception as e:
            print("Failed to read objects.json:", e)
            records = []
    return jsonify(records)

if __name__ == "__main__":
    init_db(); ensure_migrations()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)