import os
from datetime import datetime, timezone

from sqlalchemy import func

from liftcrm import config
from liftcrm.db import Asset, SessionLocal, Ticket, ensure_migrations, init_db

REQUIRED_HEADERS = {"object_name", "address", "lat", "lon"}


def normalize_address(value: str) -> str:
    return " ".join((value or "").lower().split())


def round_coord(value):
    if value in (None, ""):
        return None
    return round(float(value), 5)


def seed_assets_from_xlsx(xlsx_path: str) -> int:
    if not os.path.exists(xlsx_path):
        print(f"objects.xlsx not found at {xlsx_path}")
        return 0

    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not REQUIRED_HEADERS.issubset(set(header)):
        missing = REQUIRED_HEADERS.difference(set(header))
        print(f"Missing headers in objects.xlsx: {', '.join(sorted(missing))}")
        return 0

    header_map = {name: idx for idx, name in enumerate(header)}
    now = datetime.now(timezone.utc)
    created = 0

    with SessionLocal() as db:
        for row in rows[1:]:
            object_name = row[header_map["object_name"]] if "object_name" in header_map else None
            address = row[header_map["address"]] if "address" in header_map else None
            lat = row[header_map["lat"]] if "lat" in header_map else None
            lon = row[header_map["lon"]] if "lon" in header_map else None

            address_clean = (address or "").strip()
            if not address_clean:
                continue
            address_norm = normalize_address(address_clean)
            lat_key = round_coord(lat)
            lon_key = round_coord(lon)

            existing = None
            if lat_key is not None and lon_key is not None:
                existing = (
                    db.query(Asset)
                    .filter(func.round(Asset.lat, 5) == lat_key)
                    .filter(func.round(Asset.lon, 5) == lon_key)
                    .filter(func.lower(func.trim(Asset.address)) == address_norm)
                    .first()
                )
            else:
                existing = db.query(Asset).filter(func.lower(func.trim(Asset.address)) == address_norm).first()

            if existing:
                continue

            asset = Asset(
                address=address_clean,
                entrance=None,
                lift_label=(str(object_name).strip() if object_name else None),
                serial_no=None,
                customer_id=None,
                lat=round_coord(lat),
                lon=round_coord(lon),
                status="ACTIVE",
                created_at=now,
                updated_at=now,
            )
            db.add(asset)
            created += 1

        db.commit()

    return created


def link_tickets_to_assets() -> int:
    linked = 0
    with SessionLocal() as db:
        tickets = db.query(Ticket).filter(Ticket.asset_id.is_(None)).all()
        for ticket in tickets:
            lat_key = round_coord(ticket.lat)
            lon_key = round_coord(ticket.lon)
            asset = None
            if lat_key is not None and lon_key is not None:
                asset = (
                    db.query(Asset)
                    .filter(func.round(Asset.lat, 5) == lat_key)
                    .filter(func.round(Asset.lon, 5) == lon_key)
                    .first()
                )
            if not asset and ticket.address:
                addr = normalize_address(ticket.address)
                asset = db.query(Asset).filter(func.lower(func.trim(Asset.address)) == addr).first()
            if asset:
                ticket.asset_id = asset.id
                linked += 1
        db.commit()
    return linked


def main():
    init_db()
    ensure_migrations()
    xlsx_path = os.path.join(config.OBJECTS_DIR, "objects.xlsx")
    created = seed_assets_from_xlsx(xlsx_path)
    linked = link_tickets_to_assets()
    print(f"Seeded assets: {created}")
    print(f"Linked tickets: {linked}")


if __name__ == "__main__":
    main()
