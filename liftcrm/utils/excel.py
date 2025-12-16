import os
from openpyxl import Workbook, load_workbook


def ensure_archive_headers(path, header):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        wb.save(path)


def load_or_create_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.save(path)
    return wb
