import logging
import os
from datetime import datetime, timezone
from typing import List, Tuple

from .. import config

logger = logging.getLogger(__name__)

REQUIRED_HEADERS: List[str] = ["object_name", "address", "lat", "lon", "created_at", "updated_at"]


def ensure_objects_workbook() -> str:
    """
    Ensure objects.xlsx exists with the required headers.
    Returns the absolute path to the workbook.
    """
    os.makedirs(config.OBJECTS_DIR, exist_ok=True)
    xlsx_path = os.path.join(config.OBJECTS_DIR, "objects.xlsx")
    if not os.path.exists(xlsx_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(REQUIRED_HEADERS)
        now = datetime.now(timezone.utc).isoformat()
        ws.append(["Central Almaty", "ул. Кабанбай Батыра 123", 43.238949, 76.889709, now, now])
        wb.save(xlsx_path)
    return xlsx_path


def _normalize_header(ws) -> Tuple[List[str], bool]:
    """
    Return the header row values and whether it was modified to include missing columns.
    """
    header = []
    if ws.max_row >= 1:
        header = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
    changed = False
    if not header:
        header = list(REQUIRED_HEADERS)
        ws.append(header)
        changed = True
    else:
        for col in REQUIRED_HEADERS:
            if col not in header:
                header.append(col)
                changed = True
        if changed:
            for idx, col in enumerate(header, start=1):
                ws.cell(row=1, column=idx, value=col)
    return header, changed


def _safe_float(val):
    try:
        return float(val)
    except Exception:
        return None


def upsert_object_from_ticket(object_name, address, lat, lon, ticket_id=None):
    """
    Best-effort upsert of an object into objects.xlsx based on geo identity.
    """
    lat_val = _safe_float(lat)
    lon_val = _safe_float(lon)
    if lat_val is None or lon_val is None:
        logger.warning("objects upsert skipped: missing coords", extra={"ticket_id": ticket_id})
        return

    try:
        xlsx_path = ensure_objects_workbook()
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active
        header, _ = _normalize_header(ws)
        header_map = {name: idx for idx, name in enumerate(header)}
        lat_idx = header_map.get("lat")
        lon_idx = header_map.get("lon")
        lat_key = round(lat_val, 5)
        lon_key = round(lon_val, 5)
        match_row = None

        for row_idx in range(2, ws.max_row + 1):
            existing_lat = _safe_float(ws.cell(row=row_idx, column=lat_idx + 1).value if lat_idx is not None else None)
            existing_lon = _safe_float(ws.cell(row=row_idx, column=lon_idx + 1).value if lon_idx is not None else None)
            if existing_lat is None or existing_lon is None:
                continue
            if round(existing_lat, 5) == lat_key and round(existing_lon, 5) == lon_key:
                match_row = row_idx
                break

        now = datetime.now(timezone.utc).isoformat()
        obj_name_clean = (object_name or "").strip()
        address_clean = (address or "").strip()

        def _set(row: int, key: str, value):
            idx = header_map.get(key)
            if idx is not None:
                ws.cell(row=row, column=idx + 1, value=value)

        if match_row:
            _set(match_row, "object_name", obj_name_clean)
            _set(match_row, "address", address_clean)
            _set(match_row, "lat", lat_val)
            _set(match_row, "lon", lon_val)
            _set(match_row, "updated_at", now)
        else:
            row_data: List = [None] * len(header)
            for key, value in [
                ("object_name", obj_name_clean),
                ("address", address_clean),
                ("lat", lat_val),
                ("lon", lon_val),
                ("created_at", now),
                ("updated_at", now),
            ]:
                idx = header_map.get(key)
                if idx is not None:
                    if len(row_data) <= idx:
                        row_data.extend([None] * (idx + 1 - len(row_data)))
                    row_data[idx] = value
            ws.append(row_data)

        wb.save(xlsx_path)
    except Exception:
        logger.warning("objects upsert failed", extra={"ticket_id": ticket_id}, exc_info=True)
