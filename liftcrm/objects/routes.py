import os
from flask import Blueprint, jsonify
from flask_login import login_required
from .. import config
from .service import ensure_objects_workbook

bp = Blueprint("objects", __name__)


@bp.get("/api/objects")
@login_required
def api_objects():
    objects_dir = config.OBJECTS_DIR
    try:
        xlsx_path = ensure_objects_workbook()
    except Exception as e:
        print("Failed to ensure objects.xlsx:", e)
        xlsx_path = os.path.join(objects_dir, "objects.xlsx")
    json_path = os.path.join(objects_dir, "objects.json")
    records = []
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h) if h is not None else "" for h in rows[0]]
                for r in rows[1:]:
                    obj = {header[i]: r[i] for i in range(len(header))}
                    try:
                        lat = float(obj.get("lat")) if obj.get("lat") is not None else None
                        lon = float(obj.get("lon")) if obj.get("lon") is not None else None
                    except Exception:
                        lat = None
                        lon = None
                    if lat is None or lon is None:
                        continue
                    records.append(
                        {
                            "object_name": str(obj.get("object_name", "")),
                            "address": str(obj.get("address", "")),
                            "lat": lat,
                            "lon": lon,
                        }
                    )
        except Exception as e:
            print("Failed to read objects.xlsx:", e)
            records = []
    if not records and os.path.exists(json_path):
        try:
            import json

            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                for obj in data:
                    try:
                        lat = float(obj.get("lat")) if obj.get("lat") is not None else None
                        lon = float(obj.get("lon")) if obj.get("lon") is not None else None
                    except Exception:
                        lat = None
                        lon = None
                    if lat is None or lon is None:
                        continue
                    records.append(
                        {
                            "object_name": str(obj.get("object_name", "")),
                            "address": str(obj.get("address", "")),
                            "lat": lat,
                            "lon": lon,
                        }
                    )
        except Exception as e:
            print("Failed to read objects.json:", e)
            records = []
    return jsonify(records)
