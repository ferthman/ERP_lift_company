import os
import glob
import re
import shutil
from datetime import datetime, timezone
from math import radians, sin, cos, atan2, sqrt

from . import repository
from ..db import Master, Ticket
from ..utils.time import to_utc
from .. import config

CANCEL_REASONS = {
    "CLIENT_REQUEST",
    "DUPLICATE",
    "NO_ACCESS",
    "OUT_OF_SCOPE",
    "OTHER",
}

def haversine_m(lat1, lon1, lat2, lon2):
    R = 6371000.0
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dl = radians(lon2 - lon1)
    a = sin(dphi / 2) ** 2 + cos(phi1) * cos(phi2) * sin(dl / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return R * c


def auto_assign_master(db):
    open_statuses = ["NEW", "ASSIGNED", "IN_PROGRESS"]
    active = db.query(Master).filter(Master.is_active == 1).all()
    counts = {m.id: 0 for m in active}
    if active:
        rows = (
            db.query(Ticket.assigned_master_id, repository.func.count(Ticket.id))
            .filter(
                Ticket.status.in_(open_statuses),
                Ticket.assigned_master_id.in_([m.id for m in active]),
                Ticket.archived_at.is_(None),
            )
            .group_by(Ticket.assigned_master_id)
            .all()
        )
        for mid, cnt in rows:
            counts[mid] = cnt
    if not counts:
        return None
    best_id = min(counts, key=lambda k: (counts[k], k))
    return db.get(Master, best_id)


def send_report(ticket: Ticket):
    if not getattr(ticket, "email", None):
        return
    import smtplib
    from email.mime.text import MIMEText

    smtp_server = config.SMTP_SERVER
    smtp_port = config.SMTP_PORT
    smtp_user = config.SMTP_USERNAME
    smtp_pass = config.SMTP_PASSWORD
    if not smtp_server or not smtp_port:
        print("Email report skipped: SMTP_SERVER or SMTP_PORT not configured")
        return
    subject = f"Заявка {ticket.id} — {ticket.object_name} завершена"
    body = (
        f"Здравствуйте!\n\n"
        f"Ваша заявка '{ticket.object_name}' была завершена.\n"
        f"Статус: {ticket.status}\n"
        f"Описание: {ticket.description or ''}\n"
        f"Адрес: {ticket.address or ''}\n\n"
        f"Спасибо за обращение."
    )
    msg = MIMEText(body, _charset="utf-8")
    msg["Subject"] = subject
    msg["From"] = smtp_user or "noreply@example.com"
    msg["To"] = ticket.email
    try:
        with smtplib.SMTP(smtp_server, int(smtp_port)) as server:
            server.starttls()
            if smtp_user:
                try:
                    server.login(smtp_user, smtp_pass or "")
                except Exception:
                    pass
            server.sendmail(msg["From"], [ticket.email], msg.as_string())
    except Exception as e:
        print("Failed to send completion report:", e)


def archive_ticket(ticket: Ticket, archive_path: str):
    header = [
        "id",
        "object_name",
        "address",
        "lat",
        "lon",
        "description",
        "priority",
        "email",
        "status",
        "close_reason",
        "close_comment",
        "assigned_master_id",
        "assigned_master_name",
        "created_at",
        "updated_at",
        "arrived_at",
        "completed_at",
        "archived_at",
        "custom_sla_response_minutes",
        "custom_sla_completion_minutes",
        "sla_response_breached",
        "sla_completion_breached",
    ]
    try:
        from openpyxl import Workbook, load_workbook

        if not os.path.exists(archive_path):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            wb.save(archive_path)
            wb = load_workbook(archive_path)
        ws = wb.active
        sla = repository.compute_sla_fields(ticket)
        row_data = [
            ticket.id,
            ticket.object_name,
            ticket.address,
            ticket.lat,
            ticket.lon,
            ticket.description,
            ticket.priority or "MEDIUM",
            ticket.email,
            ticket.status,
            ticket.close_reason,
            ticket.close_comment,
            ticket.assigned_master_id,
            ticket.assigned_master.name if ticket.assigned_master else None,
            ticket.created_at.isoformat() if ticket.created_at else None,
            ticket.updated_at.isoformat() if ticket.updated_at else None,
            ticket.arrived_at.isoformat() if ticket.arrived_at else None,
            ticket.completed_at.isoformat() if ticket.completed_at else None,
            ticket.archived_at.isoformat() if ticket.archived_at else None,
            ticket.custom_sla_response_minutes,
            ticket.custom_sla_completion_minutes,
            sla.get("sla_response_breached"),
            sla.get("sla_completion_breached"),
        ]
        ws.append(row_data)
        wb.save(archive_path)
        try:
            dirpath = os.path.dirname(archive_path)
            existing = glob.glob(os.path.join(dirpath, "archive_*.xlsx"))
            nums = []
            for f in existing:
                m = re.search(r"archive_(\d+)\.xlsx", os.path.basename(f))
                if m:
                    nums.append(int(m.group(1)))
            next_num = max(nums) + 1 if nums else 1
            numbered_path = os.path.join(dirpath, f"archive_{next_num}.xlsx")
            shutil.copy2(archive_path, numbered_path)
        except Exception as e_inner:
            print("Failed to create numbered archive copy:", e_inner)
    except Exception as e:
        print("Failed to archive ticket:", e)


def _to_utc(dt):
    return to_utc(dt)


def validate_status_transition(old_status, new_status, ticket, actor_role, payload):
    allowed_transitions = {
        "NEW": {"ASSIGNED", "CANCELLED"},
        "ASSIGNED": {"IN_PROGRESS", "CANCELLED"},
        "IN_PROGRESS": {"COMPLETED", "CANCELLED"},
        "COMPLETED": {"CANCELLED"},
        "CANCELLED": set(),
    }
    if old_status == new_status:
        return _validate_status_invariants(new_status, ticket, actor_role, payload)
    allowed = allowed_transitions.get(old_status, set())
    if new_status not in allowed:
        hint = _transition_hint(old_status, new_status)
        if hint:
            message = f"Invalid transition {old_status} -> {new_status}. {hint}"
        elif allowed:
            allowed_list = "/".join(sorted(allowed))
            message = f"Invalid transition {old_status} -> {new_status}. Allowed: {allowed_list}."
        else:
            message = f"Invalid transition {old_status} -> {new_status}."
        return False, "INVALID_STATUS_TRANSITION", message
    return _validate_status_invariants(new_status, ticket, actor_role, payload)


def _transition_hint(old_status, new_status):
    hints = {
        ("NEW", "IN_PROGRESS"): "Must go through ASSIGNED.",
        ("NEW", "COMPLETED"): "Must go through ASSIGNED/IN_PROGRESS.",
        ("ASSIGNED", "COMPLETED"): "Must go through IN_PROGRESS.",
    }
    return hints.get((old_status, new_status))


def _validate_status_invariants(new_status, ticket, actor_role, payload):
    if new_status == "ASSIGNED" and not ticket.assigned_master_id:
        return (
            False,
            "INVALID_STATUS_TRANSITION",
            "Cannot set status ASSIGNED without assigned_master_id.",
        )
    if new_status == "IN_PROGRESS" and not ticket.assigned_master_id:
        return (
            False,
            "INVALID_STATUS_TRANSITION",
            "Cannot set status IN_PROGRESS without assigned_master_id.",
        )
    if new_status == "COMPLETED":
        if not ticket.completed_at:
            return (
                False,
                "INVALID_STATUS_TRANSITION",
                "Cannot set status COMPLETED without completed_at.",
            )
        if not ticket.close_reason:
            return (
                False,
                "INVALID_STATUS_TRANSITION",
                "Cannot set status COMPLETED without close_reason.",
            )
    if new_status == "CANCELLED":
        if actor_role not in {"admin", "dispatcher"}:
            return (
                False,
                "FORBIDDEN",
                "Only admin/dispatcher can cancel tickets.",
            )
        close_reason = (payload or {}).get("close_reason")
        close_comment = (payload or {}).get("close_comment")
        if not close_reason:
            return (
                False,
                "INVALID_STATUS_TRANSITION",
                "Close reason is required to cancel a ticket.",
            )
        if close_reason not in CANCEL_REASONS:
            return (
                False,
                "INVALID_STATUS_TRANSITION",
                "Invalid close_reason for cancellation.",
            )
        if not close_comment or len(str(close_comment).strip()) < 5:
            return (
                False,
                "INVALID_STATUS_TRANSITION",
                "Close comment must be at least 5 characters.",
            )
    return True, "", ""
