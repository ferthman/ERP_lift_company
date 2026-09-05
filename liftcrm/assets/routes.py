import csv
import io
import json
import logging
from datetime import date, datetime, timezone, timedelta, time

from flask import Blueprint, jsonify, request, send_file
from flask_login import current_user, login_required

from ..db import SessionLocal, Asset, Ticket, TicketComment, Attachment, AuditLog, User, Master, Customer, Contract, MaintenancePlan
from ..tickets.service import auto_assign_master
from ..utils.audit import log_audit
from ..utils.security import role_required
from ..utils.time import to_utc
from .service import normalize_text
from ..buildings.service import link_asset, coordinates

bp = Blueprint("assets", __name__)
logger = logging.getLogger(__name__)

ASSET_IMPORT_ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
ASSET_IMPORT_HEADERS = {
    "address": {
        "address",
        "object",
        "object_name",
        "объект",
        "адрес",
    },
    "entrance": {
        "entrance",
        "подъезд",
    },
    "lift_label": {
        "lift_label",
        "lift",
        "elevator",
        "elevator_label",
        "лифт",
        "метка лифта",
    },
    "serial_no": {
        "serial_no",
        "serial",
        "заводской номер",
        "серийный номер",
    },
    "lat": {
        "latitude",
        "lat",
        "широта",
    },
    "lon": {
        "longitude",
        "lng",
        "lon",
        "долгота",
    },
    "status": {
        "status",
        "статус",
    },
}
ASSET_IMPORT_ALIAS_TO_FIELD = {
    alias: field for field, aliases in ASSET_IMPORT_HEADERS.items() for alias in aliases
}
CONTRACT_STATUSES = {"active", "paused", "expired"}
MAINTENANCE_INTERVALS = {"monthly", "quarterly", "semiannual", "annual", "custom"}
MAINTENANCE_STATUSES = {"active", "paused", "completed", "overdue"}
MAINTENANCE_ACTIVE_STATUSES = {"active", "overdue"}
OPEN_TICKET_STATUSES = {"NEW", "ASSIGNED", "ACCEPTED", "IN_PROGRESS", "WAITING"}


def serialize_asset(asset: Asset):
    customer = asset.customer
    contract = asset.contract
    return {
        "id": asset.id,
        "address": asset.address,
        "building_id": asset.building_id,
        "building_name": asset.building.name if asset.building else None,
        "entrance": asset.entrance,
        "lift_label": asset.lift_label,
        "serial_no": asset.serial_no,
        "customer_id": asset.customer_id,
        "customer_name": customer.name if customer else None,
        "contract_id": asset.contract_id,
        "contract_number": contract.contract_number if contract else None,
        "contract_title": contract.title if contract else None,
        "contract_status": contract.status if contract else None,
        "lat": asset.lat,
        "lon": asset.lon,
        "status": asset.status,
        "created_at": (to_utc(asset.created_at).isoformat() if asset.created_at else None),
        "updated_at": (to_utc(asset.updated_at).isoformat() if asset.updated_at else None),
    }


def serialize_customer(customer: Customer):
    return {
        "id": customer.id,
        "name": customer.name,
        "contact_person": customer.contact_person,
        "phone": customer.phone,
        "email": customer.email,
        "notes": customer.notes,
        "is_active": bool(customer.is_active),
        "created_at": (to_utc(customer.created_at).isoformat() if customer.created_at else None),
        "updated_at": (to_utc(customer.updated_at).isoformat() if customer.updated_at else None),
    }


def serialize_contract(contract: Contract):
    return {
        "id": contract.id,
        "customer_id": contract.customer_id,
        "customer_name": contract.customer.name if contract.customer else None,
        "contract_number": contract.contract_number,
        "title": contract.title,
        "start_date": contract.start_date.isoformat() if contract.start_date else None,
        "end_date": contract.end_date.isoformat() if contract.end_date else None,
        "status": contract.status,
        "sla_hours_normal": contract.sla_hours_normal,
        "sla_hours_high": contract.sla_hours_high,
        "sla_hours_emergency": contract.sla_hours_emergency,
        "notes": contract.notes,
        "created_at": (to_utc(contract.created_at).isoformat() if contract.created_at else None),
        "updated_at": (to_utc(contract.updated_at).isoformat() if contract.updated_at else None),
    }


def _asset_label(asset):
    if not asset:
        return None
    parts = [asset.address]
    details = " / ".join([part for part in (asset.entrance, asset.lift_label, asset.serial_no) if part])
    if details:
        parts.append(details)
    return " — ".join([part for part in parts if part])


def _maintenance_due_status(plan: MaintenancePlan, today=None):
    today = today or date.today()
    status = plan.status or "active"
    if status == "active" and plan.next_due_date and plan.next_due_date < today:
        return "overdue"
    return status


def _maintenance_due_bucket(plan: MaintenancePlan, today=None):
    today = today or date.today()
    due_status = _maintenance_due_status(plan, today)
    if due_status in {"paused", "completed"}:
        return due_status
    if due_status == "overdue":
        return "overdue"
    if plan.next_due_date == today:
        return "today"
    if plan.next_due_date and plan.next_due_date <= today + timedelta(days=7):
        return "next_7_days"
    if plan.next_due_date and plan.next_due_date <= today + timedelta(days=30):
        return "next_30_days"
    return "later"


def _generated_ticket_for_due_date(db, plan):
    if not plan.id or not plan.next_due_date:
        return None
    return (
        db.query(Ticket)
        .filter(
            Ticket.maintenance_plan_id == plan.id,
            Ticket.maintenance_due_date == plan.next_due_date,
            Ticket.archived_at.is_(None),
            Ticket.status.in_(OPEN_TICKET_STATUSES),
        )
        .order_by(Ticket.id.asc())
        .first()
    )


def serialize_maintenance_plan(plan: MaintenancePlan, today=None, generated_ticket=None):
    asset = plan.asset
    master = plan.assigned_master
    customer = asset.customer if asset else None
    contract = asset.contract if asset else None
    due_status = _maintenance_due_status(plan, today)
    due_bucket = _maintenance_due_bucket(plan, today)
    return {
        "id": plan.id,
        "asset_id": plan.asset_id,
        "asset_label": _asset_label(asset),
        "asset_address": asset.address if asset else None,
        "asset_lift_label": asset.lift_label if asset else None,
        "asset_entrance": asset.entrance if asset else None,
        "customer_id": asset.customer_id if asset else None,
        "customer_name": customer.name if customer else None,
        "contract_id": asset.contract_id if asset else None,
        "contract_title": contract.title if contract else None,
        "contract_number": contract.contract_number if contract else None,
        "contract_status": contract.status if contract else None,
        "title": plan.title,
        "description": plan.description,
        "interval_type": plan.interval_type,
        "next_due_date": plan.next_due_date.isoformat() if plan.next_due_date else None,
        "last_completed_date": plan.last_completed_date.isoformat() if plan.last_completed_date else None,
        "assigned_master_id": plan.assigned_master_id,
        "assigned_master_name": master.name if master else None,
        "status": plan.status,
        "due_status": due_status,
        "due_bucket": due_bucket,
        "generated_ticket_id": generated_ticket.id if generated_ticket else None,
        "generated_ticket_status": generated_ticket.status if generated_ticket else None,
        "notes": plan.notes,
        "created_at": (to_utc(plan.created_at).isoformat() if plan.created_at else None),
        "updated_at": (to_utc(plan.updated_at).isoformat() if plan.updated_at else None),
    }


def _clean_optional(value):
    return str(value or "").strip() or None


def _parse_bool_int(value, default=True):
    if value is None:
        return 1 if default else 0
    if isinstance(value, bool):
        return 1 if value else 0
    raw = str(value).strip().lower()
    if raw in {"1", "true", "yes", "active"}:
        return 1
    if raw in {"0", "false", "no", "inactive"}:
        return 0
    raise ValueError("is_active must be true or false")


def _parse_contract_date(value, field):
    if value in (None, ""):
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field} must use YYYY-MM-DD") from exc


def _parse_required_date(value, field):
    if value in (None, ""):
        raise ValueError(f"{field} is required")
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field} must use YYYY-MM-DD") from exc


def _parse_optional_date(value, field):
    if value in (None, ""):
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field} must use YYYY-MM-DD") from exc


def _parse_bool_query(value):
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _parse_positive_float(value, field):
    if value in (None, ""):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} must be a positive number") from exc
    if parsed <= 0:
        raise ValueError(f"{field} must be a positive number")
    return parsed


def _parse_optional_int(value, field):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} must be an integer") from exc


def _add_months(value, months):
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    month_lengths = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(value.day, month_lengths[month - 1])
    return date(year, month, day)


def _next_due_after_completion(completed_date, interval_type):
    if interval_type == "monthly":
        return _add_months(completed_date, 1)
    if interval_type == "quarterly":
        return _add_months(completed_date, 3)
    if interval_type == "semiannual":
        return _add_months(completed_date, 6)
    if interval_type == "annual":
        return _add_months(completed_date, 12)
    return None


def _maintenance_values_from_payload(data, db, existing=None):
    values = {}
    if existing is None or "asset_id" in data:
        asset_id = _parse_optional_int(data.get("asset_id"), "asset_id")
        if asset_id is None:
            raise ValueError("asset_id is required")
        asset = db.get(Asset, asset_id)
        if not asset:
            raise ValueError("Asset not found")
        values["asset_id"] = asset_id
    if existing is None or "title" in data:
        title = str(data.get("title") or "").strip()
        if not title:
            raise ValueError("title is required")
        values["title"] = title
    if existing is None or "interval_type" in data:
        interval_type = str(data.get("interval_type") or "").strip().lower()
        if interval_type not in MAINTENANCE_INTERVALS:
            raise ValueError("interval_type must be monthly, quarterly, semiannual, annual, or custom")
        values["interval_type"] = interval_type
    if existing is None or "next_due_date" in data:
        values["next_due_date"] = _parse_required_date(data.get("next_due_date"), "next_due_date")
    if "last_completed_date" in data:
        value = data.get("last_completed_date")
        values["last_completed_date"] = None if value in (None, "") else _parse_required_date(value, "last_completed_date")
    if existing is None or "status" in data:
        status = str(data.get("status") or "active").strip().lower()
        if status not in MAINTENANCE_STATUSES:
            raise ValueError("status must be active, paused, completed, or overdue")
        values["status"] = status
    if existing is None or "assigned_master_id" in data:
        assigned_master_id = _parse_optional_int(data.get("assigned_master_id"), "assigned_master_id")
        if assigned_master_id is not None:
            master = db.get(Master, assigned_master_id)
            if not master:
                raise ValueError("Assigned master not found")
        values["assigned_master_id"] = assigned_master_id
    for field in ("description", "notes"):
        if existing is None or field in data:
            values[field] = _clean_optional(data.get(field))
    return values


def _validate_customer_contract_link(db, customer_id, contract_id):
    customer = None
    contract = None
    if customer_id is not None:
        customer = db.get(Customer, customer_id)
        if not customer:
            raise ValueError("Customer not found")
    if contract_id is not None:
        contract = db.get(Contract, contract_id)
        if not contract:
            raise ValueError("Contract not found")
        if customer_id is not None and contract.customer_id != customer_id:
            raise ValueError("Contract must belong to selected customer")
        if customer_id is None:
            customer_id = contract.customer_id
    return customer_id, contract_id, customer, contract


def _parse_date(value: str, label: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"Invalid {label} date format, expected YYYY-MM-DD") from exc


def _parse_iso_ts(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return to_utc(value)
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value)
        except ValueError:
            return None
        return to_utc(parsed)
    return None


def _serialize_actor(user):
    if not user:
        return None
    return {"id": user.id, "username": user.username, "role": user.role}


def _format_status_change(old_status, new_status):
    if old_status or new_status:
        return f"Статус {old_status or '—'} → {new_status or '—'}"
    return "Изменение статуса"


def _format_close_text(reason, comment):
    if reason and comment:
        return f"{reason}: {comment}"
    return reason or comment or "Заявка закрыта"


def _format_iso(dt_value):
    if not dt_value:
        return None
    return to_utc(dt_value).isoformat()


def _first_status_ts(status_events, status):
    for ts, state in status_events:
        if state == status:
            return ts
    return None


def _compute_ticket_metrics(ticket, status_events):
    created_at = to_utc(ticket.created_at) if ticket.created_at else None
    response_ts = None
    if ticket.arrived_at:
        response_ts = to_utc(ticket.arrived_at)
    elif ticket.accepted_at:
        response_ts = to_utc(ticket.accepted_at)
    else:
        response_ts = _first_status_ts(status_events, "IN_PROGRESS")

    response_seconds = None
    if created_at and response_ts and response_ts >= created_at:
        response_seconds = int((response_ts - created_at).total_seconds())

    in_progress_ts = _first_status_ts(status_events, "IN_PROGRESS")
    completed_ts = to_utc(ticket.completed_at) if ticket.completed_at else _first_status_ts(status_events, "COMPLETED")
    repair_seconds = None
    if in_progress_ts and completed_ts and completed_ts >= in_progress_ts:
        repair_seconds = int((completed_ts - in_progress_ts).total_seconds())

    downtime_seconds = None
    waiting_durations = []
    waiting_start = None
    last_ts = None
    for ts, status in status_events:
        last_ts = ts
        if status == "WAITING":
            if waiting_start is None:
                waiting_start = ts
        else:
            if waiting_start is not None:
                waiting_durations.append((ts - waiting_start).total_seconds())
                waiting_start = None
    if waiting_start is not None and last_ts is not None:
        waiting_durations.append((last_ts - waiting_start).total_seconds())
        waiting_start = None
    if waiting_durations:
        downtime_seconds = int(sum(waiting_durations))

    return {
        "response_seconds": response_seconds,
        "repair_seconds": repair_seconds,
        "downtime_seconds": downtime_seconds,
    }


def _ensure_unique_serial(db, serial_no, asset_id=None):
    if not serial_no:
        return None
    q = db.query(Asset).filter(Asset.serial_no == serial_no)
    if asset_id is not None:
        q = q.filter(Asset.id != asset_id)
    return q.first()


def _normalize_import_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _canonical_import_header(value):
    normalized = _normalize_import_header(value)
    underscored = normalized.replace(" ", "_")
    return ASSET_IMPORT_ALIAS_TO_FIELD.get(underscored) or ASSET_IMPORT_ALIAS_TO_FIELD.get(normalized)


def _is_empty_import_row(row):
    return all(str(value or "").strip() == "" for value in row.values())


def _parse_optional_float(value, field, row_number, errors):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return float(raw.replace(",", "."))
    except ValueError:
        errors.append({"row": row_number, "field": field, "message": f"{field} must be a valid number"})
        return None


def _normalize_import_status(value, row_number, errors):
    status = str(value or "").strip().upper() or "ACTIVE"
    if status not in {"ACTIVE", "INACTIVE"}:
        errors.append({"row": row_number, "field": "status", "message": "status must be ACTIVE or INACTIVE"})
        return None
    return status


def _asset_composite_key(address, entrance, lift_label):
    if not address or not entrance or not lift_label:
        return None
    return (
        normalize_text(address),
        normalize_text(entrance),
        normalize_text(lift_label),
    )


def _find_asset_by_composite(db, address, entrance, lift_label):
    key = _asset_composite_key(address, entrance, lift_label)
    if not key:
        return None
    address_norm, entrance_norm, lift_label_norm = key
    for asset in db.query(Asset).filter(Asset.address_norm == address_norm).all():
        if normalize_text(asset.entrance or "") == entrance_norm and normalize_text(asset.lift_label or "") == lift_label_norm:
            return asset
    return None


def _parse_import_rows_csv(file_storage):
    raw = file_storage.read()
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], ["Import file must contain a header row"]
    mapped_headers = [_canonical_import_header(name) for name in reader.fieldnames]
    if "address" not in mapped_headers:
        return [], ["Import file must include an address/object column"]
    rows = []
    for row_number, row in enumerate(reader, start=2):
        canonical = {}
        for original, field in zip(reader.fieldnames, mapped_headers):
            if field and field not in canonical:
                canonical[field] = row.get(original)
        rows.append((row_number, canonical))
    return rows, []


def _parse_import_rows_xlsx(file_storage):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        logger.exception("openpyxl missing", exc_info=exc)
        return [], ["openpyxl is required to read XLSX files"]
    try:
        wb = load_workbook(io.BytesIO(file_storage.read()), read_only=True, data_only=True)
    except Exception:
        logger.exception("asset_import_xlsx_parse_failed")
        return [], ["XLSX file could not be read"]
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], ["Import file must contain a header row"]
    mapped_headers = [_canonical_import_header(value) for value in header_row]
    if "address" not in mapped_headers:
        return [], ["Import file must include an address/object column"]
    rows = []
    for row_number, row in enumerate(iterator, start=2):
        canonical = {}
        for index, field in enumerate(mapped_headers):
            if field and field not in canonical:
                canonical[field] = row[index] if index < len(row) else None
        rows.append((row_number, canonical))
    return rows, []


def _parse_asset_import_file(file_storage):
    filename = (file_storage.filename or "").strip().lower()
    extension = ""
    if "." in filename:
        extension = filename[filename.rfind(".") :]
    if extension not in ASSET_IMPORT_ALLOWED_EXTENSIONS:
        return [], [f"Unsupported file type. Use .csv or .xlsx."]
    if extension == ".csv":
        try:
            return _parse_import_rows_csv(file_storage)
        except UnicodeDecodeError:
            return [], ["CSV file must be UTF-8 encoded"]
    return _parse_import_rows_xlsx(file_storage)


def _validate_import_row(row_number, row):
    errors = []
    if _is_empty_import_row(row):
        return None, []

    address = str(row.get("address") or "").strip()
    if not address:
        errors.append({"row": row_number, "field": "address", "message": "address is required"})
    lat = _parse_optional_float(row.get("lat"), "lat", row_number, errors)
    lon = _parse_optional_float(row.get("lon"), "lon", row_number, errors)
    try:
        lat, lon = coordinates(lat, lon)
    except ValueError as exc:
        errors.append({"row": row_number, "field": "lat/lon", "message": str(exc)})
    status = _normalize_import_status(row.get("status"), row_number, errors)
    if errors:
        return None, errors
    return {
        "address": address,
        "address_norm": normalize_text(address),
        "entrance": str(row.get("entrance") or "").strip() or None,
        "lift_label": str(row.get("lift_label") or "").strip() or None,
        "serial_no": str(row.get("serial_no") or "").strip() or None,
        "lat": lat,
        "lon": lon,
        "status": status,
    }, []


@bp.get("/api/customers")
@login_required
@role_required("admin", "dispatcher")
def list_customers():
    with SessionLocal() as db:
        customers = db.query(Customer).order_by(Customer.name.asc(), Customer.id.asc()).all()
        return jsonify([serialize_customer(customer) for customer in customers])


@bp.post("/api/customers")
@login_required
@role_required("admin", "dispatcher")
def create_customer():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Customer name is required"}), 400
    try:
        is_active = _parse_bool_int(data.get("is_active"), default=True)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    now = datetime.now(timezone.utc)
    with SessionLocal() as db:
        customer = Customer(
            name=name,
            contact_person=_clean_optional(data.get("contact_person")),
            phone=_clean_optional(data.get("phone")),
            email=_clean_optional(data.get("email")),
            notes=_clean_optional(data.get("notes")),
            is_active=is_active,
            created_at=now,
            updated_at=now,
        )
        db.add(customer)
        db.commit()
        db.refresh(customer)
        return jsonify(serialize_customer(customer)), 201


@bp.get("/api/customers/<int:customer_id>")
@login_required
@role_required("admin", "dispatcher")
def get_customer(customer_id):
    with SessionLocal() as db:
        customer = db.get(Customer, customer_id)
        if not customer:
            return jsonify({"error": "Customer not found"}), 404
        return jsonify(serialize_customer(customer))


@bp.patch("/api/customers/<int:customer_id>")
@login_required
@role_required("admin", "dispatcher")
def update_customer(customer_id):
    data = request.get_json() or {}
    with SessionLocal() as db:
        customer = db.get(Customer, customer_id)
        if not customer:
            return jsonify({"error": "Customer not found"}), 404
        if "name" in data:
            name = (data.get("name") or "").strip()
            if not name:
                return jsonify({"error": "Customer name is required"}), 400
            customer.name = name
        for field in ("contact_person", "phone", "email", "notes"):
            if field in data:
                setattr(customer, field, _clean_optional(data.get(field)))
        if "is_active" in data:
            try:
                customer.is_active = _parse_bool_int(data.get("is_active"), default=True)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
        customer.updated_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(customer)
        return jsonify(serialize_customer(customer))


def _contract_values_from_payload(data, db, existing=None):
    if existing is None or "customer_id" in data:
        customer_id = _parse_optional_int(data.get("customer_id"), "customer_id")
        if customer_id is None:
            raise ValueError("customer_id is required")
        customer = db.get(Customer, customer_id)
        if not customer:
            raise ValueError("Customer not found")
    else:
        customer_id = existing.customer_id

    title = None
    if existing is None or "title" in data or "name" in data:
        title = (data.get("title") or data.get("name") or "").strip()
        if not title:
            raise ValueError("Contract title is required")

    status = None
    if existing is None or "status" in data:
        status = str(data.get("status") or "active").strip().lower()
        if status not in CONTRACT_STATUSES:
            raise ValueError("Contract status must be active, paused, or expired")

    start_date = _parse_contract_date(data.get("start_date"), "start_date") if existing is None or "start_date" in data else None
    end_date = _parse_contract_date(data.get("end_date"), "end_date") if existing is None or "end_date" in data else None
    effective_start = start_date if start_date is not None or existing is None else existing.start_date
    effective_end = end_date if end_date is not None or existing is None else existing.end_date
    if effective_start and effective_end and effective_end < effective_start:
        raise ValueError("end_date must be on or after start_date")

    values = {"customer_id": customer_id}
    if title is not None:
        values["title"] = title
    if status is not None:
        values["status"] = status
    for field in ("contract_number", "notes"):
        if existing is None or field in data:
            values[field] = _clean_optional(data.get(field))
    if existing is None or "start_date" in data:
        values["start_date"] = start_date
    if existing is None or "end_date" in data:
        values["end_date"] = end_date
    for field in ("sla_hours_normal", "sla_hours_high", "sla_hours_emergency"):
        if existing is None or field in data:
            values[field] = _parse_positive_float(data.get(field), field)
    return values


@bp.get("/api/contracts")
@login_required
@role_required("admin", "dispatcher")
def list_contracts():
    customer_id = request.args.get("customer_id")
    try:
        customer_id = _parse_optional_int(customer_id, "customer_id")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with SessionLocal() as db:
        query = db.query(Contract).order_by(Contract.id.desc())
        if customer_id is not None:
            query = query.filter(Contract.customer_id == customer_id)
        return jsonify([serialize_contract(contract) for contract in query.all()])


@bp.post("/api/contracts")
@login_required
@role_required("admin", "dispatcher")
def create_contract():
    data = request.get_json() or {}
    with SessionLocal() as db:
        try:
            values = _contract_values_from_payload(data, db)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        now = datetime.now(timezone.utc)
        contract = Contract(**values, created_at=now, updated_at=now)
        db.add(contract)
        db.commit()
        db.refresh(contract)
        return jsonify(serialize_contract(contract)), 201


@bp.get("/api/contracts/<int:contract_id>")
@login_required
@role_required("admin", "dispatcher")
def get_contract(contract_id):
    with SessionLocal() as db:
        contract = db.get(Contract, contract_id)
        if not contract:
            return jsonify({"error": "Contract not found"}), 404
        return jsonify(serialize_contract(contract))


@bp.patch("/api/contracts/<int:contract_id>")
@login_required
@role_required("admin", "dispatcher")
def update_contract(contract_id):
    data = request.get_json() or {}
    with SessionLocal() as db:
        contract = db.get(Contract, contract_id)
        if not contract:
            return jsonify({"error": "Contract not found"}), 404
        try:
            values = _contract_values_from_payload(data, db, existing=contract)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        for field, value in values.items():
            setattr(contract, field, value)
        contract.updated_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(contract)
        return jsonify(serialize_contract(contract))


@bp.get("/api/maintenance-plans")
@login_required
@role_required("admin", "dispatcher")
def list_maintenance_plans():
    with SessionLocal() as db:
        query = db.query(MaintenancePlan).order_by(MaintenancePlan.next_due_date.asc(), MaintenancePlan.id.asc())
        asset_id = request.args.get("asset_id")
        status = request.args.get("status")
        try:
            parsed_asset_id = _parse_optional_int(asset_id, "asset_id")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        if parsed_asset_id is not None:
            query = query.filter(MaintenancePlan.asset_id == parsed_asset_id)
        if status:
            normalized_status = status.strip().lower()
            if normalized_status not in MAINTENANCE_STATUSES:
                return jsonify({"error": "status must be active, paused, completed, or overdue"}), 400
            query = query.filter(MaintenancePlan.status == normalized_status)
        return jsonify([serialize_maintenance_plan(plan) for plan in query.all()])


@bp.get("/api/maintenance-plans/due")
@login_required
@role_required("admin", "dispatcher")
def list_due_maintenance_plans():
    today = date.today()
    default_until = today + timedelta(days=30)
    with SessionLocal() as db:
        query = db.query(MaintenancePlan).order_by(MaintenancePlan.next_due_date.asc(), MaintenancePlan.id.asc())
        status = request.args.get("status")
        assigned_master_id = request.args.get("assigned_master_id")
        overdue_only = _parse_bool_query(request.args.get("overdue_only"))
        include_inactive = _parse_bool_query(request.args.get("include_inactive"))
        try:
            date_from = _parse_optional_date(request.args.get("date_from"), "date_from")
            date_to = _parse_optional_date(request.args.get("date_to"), "date_to")
            parsed_master_id = _parse_optional_int(assigned_master_id, "assigned_master_id")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        if date_from and date_to and date_to < date_from:
            return jsonify({"error": "date_to must be on or after date_from"}), 400
        if parsed_master_id is not None:
            query = query.filter(MaintenancePlan.assigned_master_id == parsed_master_id)
        if status:
            normalized_status = status.strip().lower()
            if normalized_status not in MAINTENANCE_STATUSES:
                return jsonify({"error": "status must be active, paused, completed, or overdue"}), 400
            if normalized_status == "overdue":
                query = query.filter(MaintenancePlan.status == "active", MaintenancePlan.next_due_date < today)
            else:
                query = query.filter(MaintenancePlan.status == normalized_status)
                if normalized_status in {"paused", "completed"}:
                    include_inactive = True
        elif not include_inactive:
            query = query.filter(MaintenancePlan.status.in_(MAINTENANCE_ACTIVE_STATUSES))
        if overdue_only:
            query = query.filter(MaintenancePlan.status == "active", MaintenancePlan.next_due_date < today)
        if date_from:
            query = query.filter(MaintenancePlan.next_due_date >= date_from)
        if date_to:
            query = query.filter(MaintenancePlan.next_due_date <= date_to)
        elif not overdue_only and not status:
            query = query.filter(MaintenancePlan.next_due_date <= default_until)

        plans = query.all()
        counters = {
            "overdue": 0,
            "today": 0,
            "next_7_days": 0,
            "next_30_days": 0,
            "paused": 0,
            "completed": 0,
        }
        active_items = []
        inactive_items = []
        for plan in plans:
            generated_ticket = _generated_ticket_for_due_date(db, plan)
            item = serialize_maintenance_plan(plan, today=today, generated_ticket=generated_ticket)
            bucket = item["due_bucket"]
            if bucket in counters:
                counters[bucket] += 1
            if bucket in {"paused", "completed"}:
                inactive_items.append(item)
            else:
                active_items.append(item)
        return jsonify(
            {
                "today": today.isoformat(),
                "range": {
                    "date_from": date_from.isoformat() if date_from else None,
                    "date_to": date_to.isoformat() if date_to else (None if status or overdue_only else default_until.isoformat()),
                },
                "counters": counters,
                "plans": active_items,
                "inactive_plans": inactive_items,
            }
        )


@bp.post("/api/maintenance-plans")
@login_required
@role_required("admin", "dispatcher")
def create_maintenance_plan():
    data = request.get_json() or {}
    with SessionLocal() as db:
        try:
            values = _maintenance_values_from_payload(data, db)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        now = datetime.now(timezone.utc)
        plan = MaintenancePlan(**values, created_at=now, updated_at=now)
        db.add(plan)
        db.commit()
        db.refresh(plan)
        return jsonify(serialize_maintenance_plan(plan)), 201


@bp.get("/api/maintenance-plans/<int:plan_id>")
@login_required
@role_required("admin", "dispatcher")
def get_maintenance_plan(plan_id):
    with SessionLocal() as db:
        plan = db.get(MaintenancePlan, plan_id)
        if not plan:
            return jsonify({"error": "Maintenance plan not found"}), 404
        return jsonify(serialize_maintenance_plan(plan))


@bp.patch("/api/maintenance-plans/<int:plan_id>")
@login_required
@role_required("admin", "dispatcher")
def update_maintenance_plan(plan_id):
    data = request.get_json() or {}
    with SessionLocal() as db:
        plan = db.get(MaintenancePlan, plan_id)
        if not plan:
            return jsonify({"error": "Maintenance plan not found"}), 404
        try:
            values = _maintenance_values_from_payload(data, db, existing=plan)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        for field, value in values.items():
            setattr(plan, field, value)
        plan.updated_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(plan)
        return jsonify(serialize_maintenance_plan(plan))


@bp.post("/api/maintenance-plans/<int:plan_id>/generate-ticket")
@login_required
@role_required("admin", "dispatcher")
def generate_ticket_from_maintenance_plan(plan_id):
    with SessionLocal() as db:
        plan = db.get(MaintenancePlan, plan_id)
        if not plan:
            return jsonify({"error": "Maintenance plan not found"}), 404
        if plan.status not in MAINTENANCE_ACTIVE_STATUSES:
            return jsonify({"error": "Only active maintenance plans can generate tickets"}), 400
        asset = plan.asset
        if not asset:
            return jsonify({"error": "Maintenance plan asset not found"}), 400
        existing = _generated_ticket_for_due_date(db, plan)
        if existing:
            return jsonify(
                {
                    "ticket_id": existing.id,
                    "status": existing.status,
                    "duplicate": True,
                    "plan": serialize_maintenance_plan(plan, generated_ticket=existing),
                }
            )
        if asset.lat is None or asset.lon is None:
            return jsonify({"error": "Asset must have coordinates before generating a ticket"}), 400
        due_date = plan.next_due_date
        description_parts = [
            f"Плановое ТО: {plan.title}",
            f"Дата ТО: {due_date.isoformat() if due_date else 'не указана'}",
        ]
        if plan.description:
            description_parts.append(plan.description)
        if plan.notes:
            description_parts.append(f"Заметки: {plan.notes}")
        ticket = Ticket(
            object_name=asset.lift_label or asset.serial_no or asset.address or f"Лифт #{asset.id}",
            address=asset.address,
            lat=asset.lat,
            lon=asset.lon,
            description="\n".join(description_parts),
            priority="MEDIUM",
            status="NEW",
            asset_id=asset.id,
            building_id=link_asset(db, asset).id,
            assigned_master_id=plan.assigned_master_id,
            maintenance_plan_id=plan.id,
            maintenance_due_date=due_date,
        )
        if ticket.assigned_master_id:
            ticket.status = "ASSIGNED"
            ticket.assigned_at = datetime.now(timezone.utc)
        else:
            master = auto_assign_master(db)
            if master:
                ticket.assigned_master_id = master.id
                ticket.status = "ASSIGNED"
                ticket.assigned_at = datetime.now(timezone.utc)
        db.add(ticket)
        db.flush()
        log_audit(
            db,
            entity_type="ticket",
            entity_id=ticket.id,
            action="CREATE",
            actor_user_id=current_user.id,
            old={},
            new={
                "object_name": ticket.object_name,
                "address": ticket.address,
                "priority": ticket.priority,
                "description": ticket.description,
                "asset_id": ticket.asset_id,
                "assigned_master_id": ticket.assigned_master_id,
                "status": ticket.status,
                "maintenance_plan_id": ticket.maintenance_plan_id,
                "maintenance_due_date": ticket.maintenance_due_date.isoformat() if ticket.maintenance_due_date else None,
            },
        )
        db.commit()
        db.refresh(ticket)
        return jsonify({"ticket_id": ticket.id, "status": ticket.status, "duplicate": False}), 201


@bp.post("/api/maintenance-plans/<int:plan_id>/complete")
@login_required
@role_required("admin", "dispatcher")
def complete_maintenance_plan(plan_id):
    data = request.get_json() or {}
    with SessionLocal() as db:
        plan = db.get(MaintenancePlan, plan_id)
        if not plan:
            return jsonify({"error": "Maintenance plan not found"}), 404
        try:
            completed_date = _parse_required_date(data.get("completed_date"), "completed_date") if data.get("completed_date") else date.today()
            next_due_date = None
            if data.get("next_due_date"):
                next_due_date = _parse_required_date(data.get("next_due_date"), "next_due_date")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        plan.last_completed_date = completed_date
        calculated_next_due = next_due_date or _next_due_after_completion(completed_date, plan.interval_type)
        if calculated_next_due:
            plan.next_due_date = calculated_next_due
            plan.status = "active"
        else:
            plan.status = "completed"
        if "notes" in data:
            plan.notes = _clean_optional(data.get("notes"))
        plan.updated_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(plan)
        return jsonify(serialize_maintenance_plan(plan))


@bp.get("/api/assets")
@login_required
def list_assets():
    search = request.args.get("search")
    with SessionLocal() as db:
        query = db.query(Asset)
        if current_user.role == 'technician':
            if not current_user.master_id: return jsonify([])
            query = query.filter(Asset.id.in_(db.query(Ticket.asset_id).filter(Ticket.assigned_master_id==current_user.master_id, Ticket.archived_at.is_(None))))
        assets = query.order_by(Asset.id.desc()).all()
        if search:
            term = normalize_text(search)
            if term:
                filtered = []
                for asset in assets:
                    haystack = normalize_text(
                        f"{asset.address_norm or asset.address or ''} {asset.serial_no or ''} {asset.lift_label or ''} {asset.entrance or ''}"
                    )
                    if term in haystack:
                        filtered.append(asset)
                assets = filtered
        return jsonify([serialize_asset(a) for a in assets])


@bp.post("/api/assets")
@login_required
@role_required("admin", "dispatcher")
def create_asset():
    data = request.get_json() or {}
    address = (data.get("address") or "").strip()
    if not address:
        return jsonify({"error": "Address is required"}), 400
    serial_no = (data.get("serial_no") or "").strip() or None
    try:
        customer_id = _parse_optional_int(data.get("customer_id"), "customer_id")
        contract_id = _parse_optional_int(data.get("contract_id"), "contract_id")
        lat = float(data["lat"]) if data.get("lat") not in (None, "") else None
        lon = float(data["lon"]) if data.get("lon") not in (None, "") else None
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with SessionLocal() as db:
        if serial_no and _ensure_unique_serial(db, serial_no):
            return jsonify({"error": "serial_no must be unique"}), 400
        try:
            customer_id, contract_id, _, _ = _validate_customer_contract_link(db, customer_id, contract_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        asset = Asset(
            address=address,
            address_norm=normalize_text(address),
            entrance=(data.get("entrance") or "").strip() or None,
            lift_label=(data.get("lift_label") or "").strip() or None,
            serial_no=serial_no,
            customer_id=customer_id,
            contract_id=contract_id,
            lat=lat,
            lon=lon,
            status=(data.get("status") or "ACTIVE").strip().upper(),
        )
        if asset.status not in {"ACTIVE", "INACTIVE"}:
            return jsonify({"error": "Invalid status"}), 400
        db.add(asset)
        try:
            asset.lat, asset.lon = coordinates(asset.lat, asset.lon)
            link_asset(db, asset, data.get("building_id"))
        except ValueError as exc:
            return jsonify(error=str(exc)), 400
        db.commit()
        db.refresh(asset)
        return jsonify(serialize_asset(asset)), 201


@bp.get("/api/assets/<int:asset_id>")
@login_required
@role_required("admin", "dispatcher")
def get_asset(asset_id):
    with SessionLocal() as db:
        asset = db.get(Asset, asset_id)
        if not asset:
            return jsonify({"error": "Asset not found"}), 404
        return jsonify(serialize_asset(asset))


@bp.patch("/api/assets/<int:asset_id>")
@login_required
@role_required("admin", "dispatcher")
def update_asset(asset_id):
    data = request.get_json() or {}
    try:
        new_customer_id = _parse_optional_int(data.get("customer_id"), "customer_id") if "customer_id" in data else None
        new_contract_id = _parse_optional_int(data.get("contract_id"), "contract_id") if "contract_id" in data else None
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    with SessionLocal() as db:
        asset = db.get(Asset, asset_id)
        if not asset:
            return jsonify({"error": "Asset not found"}), 404
        if "address" in data:
            address = (data.get("address") or "").strip()
            if not address:
                return jsonify({"error": "Address is required"}), 400
            asset.address = address
            asset.address_norm = normalize_text(address)
        if "entrance" in data:
            asset.entrance = (data.get("entrance") or "").strip() or None
        if "lift_label" in data:
            asset.lift_label = (data.get("lift_label") or "").strip() or None
        if "serial_no" in data:
            serial_no = (data.get("serial_no") or "").strip() or None
            if serial_no and _ensure_unique_serial(db, serial_no, asset_id=asset_id):
                return jsonify({"error": "serial_no must be unique"}), 400
            asset.serial_no = serial_no
        if "lat" in data:
            try:
                asset.lat = float(data["lat"]) if data.get("lat") not in (None, "") else None
            except ValueError:
                return jsonify({"error": "lat must be a valid number"}), 400
        if "lon" in data:
            try:
                asset.lon = float(data["lon"]) if data.get("lon") not in (None, "") else None
            except ValueError:
                return jsonify({"error": "lon must be a valid number"}), 400
        if "status" in data:
            status = (data.get("status") or "").strip().upper()
            if status not in {"ACTIVE", "INACTIVE"}:
                return jsonify({"error": "Invalid status"}), 400
            asset.status = status
        if "customer_id" in data or "contract_id" in data:
            customer_id = new_customer_id if "customer_id" in data else asset.customer_id
            contract_id = new_contract_id if "contract_id" in data else asset.contract_id
            try:
                customer_id, contract_id, _, _ = _validate_customer_contract_link(db, customer_id, contract_id)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
            asset.customer_id = customer_id
            asset.contract_id = contract_id
        try:
            asset.lat, asset.lon = coordinates(asset.lat, asset.lon)
            link_asset(db, asset, data.get("building_id"), regroup=any(k in data for k in ("address","customer_id","contract_id")))
        except ValueError as exc:
            return jsonify(error=str(exc)), 400
        asset.updated_at = datetime.now(timezone.utc)
        db.commit()
        db.refresh(asset)
        return jsonify(serialize_asset(asset))


@bp.delete("/api/assets/<int:asset_id>")
@login_required
@role_required("admin", "dispatcher")
def delete_asset(asset_id):
    with SessionLocal() as db:
        asset = db.get(Asset, asset_id)
        if not asset:
            return jsonify({"error": "Asset not found"}), 404
        asset.status = "INACTIVE"
        db.commit()
        return jsonify({"ok": True})


@bp.post("/api/assets/import")
@login_required
@role_required("admin", "dispatcher")
def import_assets():
    file_storage = request.files.get("file")
    if not file_storage or not file_storage.filename:
        return jsonify({"error": {"code": 400, "message": "Import file is required"}}), 400

    parsed_rows, parse_errors = _parse_asset_import_file(file_storage)
    if parse_errors:
        return jsonify({"error": {"code": 400, "message": parse_errors[0]}, "errors": parse_errors}), 400

    result = {
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "skipped_duplicates": 0,
        "invalid": 0,
        "errors": [],
    }
    validated_rows = []
    for row_number, row in parsed_rows:
        values, row_errors = _validate_import_row(row_number, row)
        if values is None:
            if row_errors:
                result["invalid"] += 1
                result["errors"].extend(row_errors)
            else:
                result["skipped"] += 1
            continue
        validated_rows.append((row_number, values))

    now = datetime.now(timezone.utc)
    with SessionLocal() as db:
        seen_serials = set()
        seen_composites = set()
        for row_number, values in validated_rows:
            existing = None
            serial_key = values["serial_no"]
            composite_key = _asset_composite_key(
                values["address"],
                values["entrance"],
                values["lift_label"],
            )
            if serial_key and serial_key in seen_serials:
                existing = True
            elif composite_key and composite_key in seen_composites:
                existing = True
            elif serial_key:
                existing = _ensure_unique_serial(db, serial_key)
            if existing is None and composite_key:
                existing = _find_asset_by_composite(
                    db,
                    values["address"],
                    values["entrance"],
                    values["lift_label"],
                )
            if existing is not None:
                result["skipped"] += 1
                result["skipped_duplicates"] += 1
                continue

            asset = Asset(
                address=values["address"],
                address_norm=values["address_norm"],
                entrance=values["entrance"],
                lift_label=values["lift_label"],
                serial_no=values["serial_no"],
                lat=values["lat"],
                lon=values["lon"],
                status=values["status"],
                created_at=now,
                updated_at=now,
            )
            db.add(asset)
            link_asset(db, asset)
            if serial_key:
                seen_serials.add(serial_key)
            if composite_key:
                seen_composites.add(composite_key)
            result["created"] += 1
        db.commit()

    return jsonify(result), 200


@bp.get("/api/lifts/<int:asset_id>/history")
@login_required
@role_required("admin", "dispatcher")
def lift_history(asset_id):
    q = normalize_text(request.args.get("q") or "")
    from_param = request.args.get("from")
    to_param = request.args.get("to")
    try:
        from_date = _parse_date(from_param, "from")
        to_date = _parse_date(to_param, "to")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    now = datetime.now(timezone.utc)
    if from_date is None and to_date is None:
        from_dt = now - timedelta(days=90)
        to_dt = now
    else:
        to_dt = (
            datetime.combine(to_date, time.max).replace(tzinfo=timezone.utc)
            if to_date
            else now
        )
        if from_date:
            from_dt = datetime.combine(from_date, time.min).replace(tzinfo=timezone.utc)
        else:
            from_dt = (to_dt - timedelta(days=90)).replace(hour=0, minute=0, second=0, microsecond=0)

    with SessionLocal() as db:
        asset = db.get(Asset, asset_id)
        if not asset:
            return jsonify({"error": "Asset not found"}), 404
        tickets = db.query(Ticket).filter(Ticket.asset_id == asset_id).all()
        ticket_ids = [t.id for t in tickets]
        audits = []
        comments = []
        attachments = []
        if ticket_ids:
            audits = (
                db.query(AuditLog)
                .filter(AuditLog.entity_type == "ticket", AuditLog.entity_id.in_(ticket_ids))
                .order_by(AuditLog.created_at.desc())
                .all()
            )
            comments = db.query(TicketComment).filter(TicketComment.ticket_id.in_(ticket_ids)).all()
            attachments = db.query(Attachment).filter(Attachment.ticket_id.in_(ticket_ids)).all()

        user_ids = {entry.actor_user_id for entry in audits if entry.actor_user_id}
        user_ids.update({c.user_id for c in comments if c.user_id})
        users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
        users_by_id = {u.id: u for u in users}

        master_ids = {t.assigned_master_id for t in tickets if t.assigned_master_id}
        masters = db.query(Master).filter(Master.id.in_(master_ids)).all() if master_ids else []
        masters_by_id = {m.id: m for m in masters}
        master_users = db.query(User).filter(User.master_id.in_(master_ids)).all() if master_ids else []
        users_by_master_id = {u.master_id: u for u in master_users if u.master_id}

        events_by_ticket = {t.id: [] for t in tickets}
        status_events_by_ticket = {t.id: [] for t in tickets}
        last_ts_by_ticket = {t.id: to_utc(t.created_at) if t.created_at else None for t in tickets}
        ticket_map = {t.id: t for t in tickets}

        def add_event(ts_dt, kind, actor_id, ticket_id, text, meta=None):
            if ts_dt is None:
                return
            ticket_events = events_by_ticket.get(ticket_id)
            if ticket_events is None:
                return
            ticket_events.append(
                {
                    "ts": ts_dt.isoformat(),
                    "kind": kind,
                    "actor": _serialize_actor(users_by_id.get(actor_id)),
                    "text": text,
                    "meta": meta or {},
                }
            )
            last_ts = last_ts_by_ticket.get(ticket_id)
            if last_ts is None or ts_dt > last_ts:
                last_ts_by_ticket[ticket_id] = ts_dt

        for entry in audits:
            ts_dt = _parse_iso_ts(entry.created_at)
            if ts_dt is None:
                continue
            try:
                payload = json.loads(entry.diff_json or "{}")
            except json.JSONDecodeError:
                payload = {}
            old = payload.get("old") or {}
            new = payload.get("new") or {}
            action = entry.action
            text = None
            kind = None
            meta = {"action": action}
            if action == "CREATE":
                kind = "CREATE"
                text = "Создана заявка"
                status_events_by_ticket.get(entry.entity_id, []).append((ts_dt, new.get("status") or "NEW"))
            elif action == "ASSIGN":
                kind = "ASSIGN"
                master_id = new.get("assigned_master_id")
                master = masters_by_id.get(master_id)
                if master:
                    text = f"Назначен мастер: {master.name}"
                else:
                    text = "Назначен мастер"
            elif action == "CANCEL":
                kind = "CLOSE"
                text = _format_close_text(new.get("close_reason"), new.get("close_comment"))
                status_events_by_ticket.get(entry.entity_id, []).append((ts_dt, "CANCELLED"))
            elif action == "STATUS_CHANGE":
                new_status = new.get("status")
                old_status = old.get("status")
                meta.update({"status": new_status, "old_status": old_status})
                if new_status:
                    status_events_by_ticket.get(entry.entity_id, []).append((ts_dt, new_status))
                if new_status == "WAITING" or new.get("waiting_reason"):
                    kind = "WAITING"
                    text = new.get("waiting_reason") or _format_status_change(old_status, new_status)
                elif new_status == "COMPLETED":
                    kind = "CLOSE"
                    text = _format_close_text(new.get("close_reason"), new.get("close_comment"))
                else:
                    kind = "STATUS_CHANGE"
                    text = _format_status_change(old_status, new_status)
            elif action == "EDIT":
                kind = "STATUS_CHANGE"
                changed = ", ".join(new.keys()) if isinstance(new, dict) else ""
                text = f"Обновлены поля: {changed}" if changed else "Обновление заявки"
            if kind:
                add_event(ts_dt, kind, entry.actor_user_id, entry.entity_id, text, meta=meta)

        audited_create = {entry.entity_id for entry in audits if entry.action == "CREATE"}
        for ticket in tickets:
            if ticket.id in audited_create:
                continue
            add_event(
                to_utc(ticket.created_at),
                "CREATE",
                None,
                ticket.id,
                "Создана заявка",
            )

        for comment in comments:
            add_event(
                to_utc(comment.created_at),
                "COMMENT",
                comment.user_id,
                comment.ticket_id,
                comment.body,
            )

        for attachment in attachments:
            add_event(
                to_utc(attachment.created_at),
                "ATTACHMENT",
                None,
                attachment.ticket_id,
                attachment.orig_name,
                meta={"filename": attachment.filename, "id": attachment.id},
            )

        def matches_query(ticket, ticket_events):
            if not q:
                return True
            haystack = normalize_text(
                f"{ticket.object_name or ''} {ticket.description or ''}"
            )
            if q in haystack:
                return True
            for ev in ticket_events:
                if q in normalize_text(ev.get("text") or ""):
                    return True
            return False

        tickets_payload = []
        for ticket in tickets:
            created_at = to_utc(ticket.created_at) if ticket.created_at else None
            if created_at and (created_at < from_dt or created_at > to_dt):
                continue
            ticket_events = events_by_ticket.get(ticket.id, [])
            if not matches_query(ticket, ticket_events):
                continue
            status_events = sorted(
                status_events_by_ticket.get(ticket.id, []),
                key=lambda x: x[0] or datetime.min.replace(tzinfo=timezone.utc),
            )
            metrics = _compute_ticket_metrics(ticket, status_events)
            last_ts = last_ts_by_ticket.get(ticket.id) or created_at
            master = masters_by_id.get(ticket.assigned_master_id)
            master_user = users_by_master_id.get(ticket.assigned_master_id)
            assigned = None
            if master or master_user:
                assigned = {
                    "id": master.id if master else None,
                    "name": master.name if master else None,
                    "username": master_user.username if master_user else None,
                }
            ticket_events_sorted = sorted(
                ticket_events,
                key=lambda x: x["ts"],
                reverse=True,
            )
            tickets_payload.append(
                {
                    "ticket": {
                        "id": ticket.id,
                        "title": ticket.object_name,
                        "description": ticket.description,
                        "status": ticket.status,
                        "created_at": _format_iso(ticket.created_at),
                        "completed_at": _format_iso(ticket.completed_at),
                        "accepted_at": _format_iso(ticket.accepted_at),
                        "arrived_at": _format_iso(ticket.arrived_at),
                        "waiting_at": _format_iso(ticket.waiting_at),
                        "waiting_reason": ticket.waiting_reason,
                        "assigned": assigned,
                    },
                    "events": ticket_events_sorted,
                    "summary": {
                        "last_ts": last_ts.isoformat() if last_ts else None,
                        "events_count": len(ticket_events_sorted),
                        "metrics": metrics,
                    },
                }
            )

        tickets_payload.sort(
            key=lambda entry: entry["summary"]["last_ts"] or "",
            reverse=True,
        )

        return jsonify({"lift": serialize_asset(asset), "tickets": tickets_payload})


@bp.get("/api/assets/export.xlsx")
@login_required
@role_required("admin", "dispatcher")
def export_assets_xlsx():
    with SessionLocal() as db:
        rows = db.query(Asset).order_by(Asset.id).all()
    try:
        from openpyxl import Workbook
    except Exception as exc:
        logger.exception("openpyxl missing", exc_info=exc)
        return jsonify({"error": "openpyxl is required"}), 500
    wb = Workbook()
    ws = wb.active
    headers = [
        "id",
        "address",
        "entrance",
        "lift_label",
        "serial_no",
        "lat",
        "lon",
        "status",
        "created_at",
        "updated_at",
    ]
    ws.append(headers)
    for a in rows:
        ws.append(
            [
                a.id,
                a.address,
                a.entrance,
                a.lift_label,
                a.serial_no,
                a.lat,
                a.lon,
                a.status,
                a.created_at.isoformat() if a.created_at else None,
                a.updated_at.isoformat() if a.updated_at else None,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="assets.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/api/assets/export.csv")
@login_required
@role_required("admin", "dispatcher")
def export_assets_csv():
    with SessionLocal() as db:
        rows = db.query(Asset).order_by(Asset.id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    headers = [
        "id",
        "address",
        "entrance",
        "lift_label",
        "serial_no",
        "lat",
        "lon",
        "status",
        "created_at",
        "updated_at",
    ]
    writer.writerow(headers)
    for a in rows:
        writer.writerow(
            [
                a.id,
                a.address,
                a.entrance,
                a.lift_label,
                a.serial_no,
                a.lat,
                a.lon,
                a.status,
                a.created_at.isoformat() if a.created_at else None,
                a.updated_at.isoformat() if a.updated_at else None,
            ]
        )
    buf = io.BytesIO(output.getvalue().encode("utf-8"))
    return send_file(
        buf,
        as_attachment=True,
        download_name="assets.csv",
        mimetype="text/csv",
    )
